import sys
import imutils
import cv2
import os
from pyimagesearch.face_blurring import anonymize_face_pixelate
from pyimagesearch.face_blurring import anonymize_face_simple
from PIL import Image, ImageOps
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5 import uic, QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from datetime import datetime
from openpyxl import Workbook, workbook
import numpy as np
from openpyxl.reader.excel import load_workbook

calUI = './FIXED_ROI.ui'
buttonPush_count = 0


cam_sub_points_by_top = {
    1:[0,0,0,0],
    2:[5,135,-3,14],
    3:[22,137,7,5],
    4:[5,-200,2,-1],
    5:[0,0,0,0],
    6:[22,98,-19,-15],
    7:[-47,-17,-16,-13],
    8:[22,-173, -11, -11],
    9:[0,0,0,0],
    10:[-32,106,15,-3],
    11:[-44,-164,5,-9],
    12:[-44,-374,8,-10],
    25:[0,0,0,0],
    26:[-44,61,-3,0],
    27:[-32,-71,-2,5],
    28:[-17,-113,5,-14],
    29:[0,0,0,0],
    30:[7,36,-12,8],
    31:[-44,-54,-8,17],
    32:[-15,-262,-14,14]
}
cam_points_by_top = {}

class MainDialog(QDialog):
    ROI = []
    def __init__(self):
        QDialog.__init__(self, None)
        self.msg = QMessageBox()
        self.msg.setWindowTitle("EXIT")
        self.msg.setText("End of File , Open New Folder")
        uic.loadUi(calUI, self)
        self.OPEN.clicked.connect(lambda: self.image_list())
        self.listWidget.itemDoubleClicked.connect(lambda: self.dragImage())
        self.listWidget.keyPressEvent = self.listWidget_keyPressEvent
        #self.SAVE.clicked.connect(lambda: self.image_save())
        #self.BLUR.clicked.connect(lambda: self.test_blurring())  
        #self.LOAD.clicked.connect(lambda: self.load_to_frame())
        self.AUTO.clicked.connect(lambda: self.auto_progress())

# OPEN
    def listWidget_keyPressEvent(self, event):
        QListWidget.keyPressEvent(self.listWidget, event)

        # user new event
        if event.key() in (Qt.Key_Return, Qt.Key_Delete):    # delete tests
            self.dragImage()
        else:
            pass

    def image_list(self):
        self.listWidget.clear()
        self.AUTO.setEnabled(True)

        dir_Name = QFileDialog.getExistingDirectory()
        if dir_Name:
            self.filepath = dir_Name
            filenames = os.listdir(dir_Name)
            self.count = 0
            filenames = [file for file in filenames if file.endswith((".png", ".PNG", ".JPG", ".jpg"))]  
            for filename in filenames:
                self.listWidget.insertItem(self.count, filename)  
                self.count += 1
        else:
            self.msg.setText("Please try again. Image list not found")
            self.msg.exec_()
            self.image_list()

    def setPhoto(self, image):
        self.tmp = image
        image = imutils.resize(image, width=640)
        frame = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image = QImage(frame, frame.shape[1], frame.shape[0], QImage.Format_RGB888)
        self.label.setPixmap(QtGui.QPixmap.fromImage(image))
        
    def image_save(self):
        dir_name = datetime.today().strftime("%Y%m%d") + "_autopass"
        model_name = str(self.filename).split("_")[2]
        save_path = os.getcwd() + "\\{}\\{}\\".format(dir_name, model_name)
        if os.path.exists(save_path):
            print(self.filename[0:-4] + ' : 저장되었습니다.')
            save_filename = self.filename[0:-4] + '.jpg'
            path = save_path + save_filename
            cv2.imwrite(path, self.image)

            output_img = Image.open(path)
            exif_dict = self.ex_img.info.get('exif')

            if exif_dict:
                EXIF = self.ex_img.info['exif']
                output_img = output_img.transpose(Image.ROTATE_90)
                output_img.save(path, exif=EXIF, dpi=(300, 300))
            else:
                output_img.save(path, dpi=(300, 300))
        else:
            os.makedirs(save_path)
            print(self.filename[0:-4] + ' : 저장되었습니다.')
            save_filename = self.filename[0:-4] + '.jpg'
            path = save_path + save_filename
            cv2.imwrite(path, self.image)

            output_img = Image.open(path)
            exif_dict = self.ex_img.info.get('exif')

            if exif_dict:
                EXIF = self.ex_img.info['exif']
                output_img = output_img.transpose(Image.ROTATE_90)
                output_img.save(path, exif=EXIF, dpi=(300, 300))
            else:
                output_img.save(path, dpi=(300, 300))

# AUTO
    def auto_progress(self):
        global buttonPush_count
        self.AUTO.setAutoRepeat(True)
        self.AUTO.animateClick()
        self.testburring(buttonPush_count)
        buttonPush_count += 1
        if buttonPush_count == self.count:
            self.AUTO.setEnabled(False)
            QTimer.singleShot(1000, lambda: self.AUTO.setDisabled(False))
            self.msg.exec_()
            self.listWidget.clear()
            buttonPush_count = 0
            return

    def set_cam_point(self, cam_num, top_cam_points):
        for i in range(1,4):
            if (cam_num+i) not in cam_sub_points_by_top:
                continue
            diff = cam_sub_points_by_top[cam_num+i]
            points = []
            zip_object =zip(top_cam_points, diff)
            for list1, list2 in zip_object:
                points.append(list1+list2)
            cam_points_by_top[cam_num+i] = points

# DoubleClick
    def dragImage(self):
        # try:
            index = self.listWidget.currentRow()
            self.filename = self.listWidget.item(index).text()
            
            self.current_image.setText(self.filename)
            cam_num = int(os.path.splitext(self.filename)[0].split('_')[-1])
            path = os.path.join(self.filepath, self.filename).replace('/', '\\')
            img_array = np.fromfile(path, np.uint8)
            self.image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

            self.ex_img = Image.open(path)
            self.setPhoto(self.image)
            ksize = 100
            print('dragImage...')
            while True:
                cv2.namedWindow("title", cv2.WINDOW_NORMAL)
                cv2.resizeWindow("title", 800, 1200)
                x, y, w, h = cv2.selectROI("title", self.image, fromCenter=False, showCrosshair=True)

                print(x,y,w,h)
                if int(cam_num) % 4 == 1:
                    if x > 0 and y > 0:
                        self.set_cam_point(cam_num, [x,y,w,h])
                        print(cam_points_by_top)
                #만약 블러 상자를 그렸으면 cam point는 덮어씌우기
                if w > 0:
                    cam_points_by_top[cam_num] = [x,y,w,h]
                if cam_num in cam_points_by_top:
                    cam = cam_points_by_top[cam_num]
                    x = cam[0]
                    y = cam[1]
                    w = cam[2]
                    h = cam[3]
                    self.save_roi_value(x, y, w, h)
                    roi = self.image[y:y + h, x:x + w]
                    # roi = cv2.blur(roi, (ksize, ksize))
                    roi = anonymize_face_pixelate(roi, blocks=5)
                    self.image[y:y + h, x:x + w] = roi
                    break

                if w > 0 and h > 0:
                    self.save_roi_value(x,y,w,h)
                    roi = self.image[y:y + h, x:x + w]
                    # roi = cv2.blur(roi, (ksize, ksize))
                    roi = anonymize_face_pixelate(roi, blocks=5)
                    self.image[y:y + h, x:x + w] = roi
                else:
                    break
            self.setPhoto(self.image)
            self.image_save()
            cv2.destroyAllWindows()
        # except Exception as error:
        #     print(error)

    def save_roi_value(self,x,y,w,h):
        POSE_num,CLOTH_num,MODEL_num,CAM_num= self.filename.split('_')
        CAM_num = CAM_num[0:2]
        if int(CLOTH_num) == 0:
            print("do not save cloth number 0")
            return
        
        xlfile = load_workbook('./test.xlsx',data_only=True)
        
        if MODEL_num not in xlfile.sheetnames:
            write_sheet = xlfile.create_sheet(MODEL_num)
        else:
            write_sheet = xlfile[MODEL_num]

        ROI_val = str(x) +','+ str(y) + ','+ str(w) + ','+ str(h)
        write_sheet.cell(int(CAM_num),int(CLOTH_num),ROI_val)
        xlfile.save('./test.xlsx')
        
    def load_roi_value(self):
        POSE_num,CLOTH_num,MODEL_num,CAM_num= self.filename.split('_')
        if int(CLOTH_num) == 0:
            print("can't load cloth number 0")
            return 0

        CAM_num = CAM_num[0:2]
        xlfile = load_workbook('./test.xlsx',data_only=True)

        if MODEL_num not in xlfile.sheetnames:
            print("doesn't exist sheet")
            return None

        load_sheet = xlfile[MODEL_num]
        load_roi = load_sheet.cell(int(CAM_num),int(CLOTH_num)).value
        if load_roi is None:
            return None

        ROI = str(load_roi).split(',')
        ## str to int
        ROI = [int(value) for value in ROI]
        return ROI

# LOAD
    def load_to_frame(self):
        index = self.listWidget.currentRow()
        self.filename = self.listWidget.item(index).text()
        
        self.current_image.setText(self.filename)
        path = os.path.join(self.filepath, self.filename).replace('/', '\\')
        img_array = np.fromfile(path, np.uint8)
        self.image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

        self.ex_img = Image.open(path)
        self.setPhoto(self.image)
        ksize = 100

        ROI = self.load_roi_value()
        print(ROI)
        roi = self.image[ROI[1]:ROI[1] + ROI[3], ROI[0]:ROI[0] + ROI[2]] 
        roi = cv2.blur(roi, (ksize, ksize))  
        self.image[ROI[1]:ROI[1] + ROI[3], ROI[0]:ROI[0] + ROI[2]] = roi 
        self.setPhoto(self.image)
        ROI.clear()

    def testburring(self,i):
        if i+1 > self.count:
            print("end of file")
            return
        self.filename = self.listWidget.item(i).text()
        self.current_image.setText(self.filename)

        
        path = os.path.join(self.filepath, self.filename).replace('/', '\\')
        img_array = np.fromfile(path, np.uint8)
        self.image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
        self.ex_img = Image.open(path)

        
        ROI = self.load_roi_value()
        
        if ROI is None:
            print("can't load ROI")
            self.image_save()
            return
        elif ROI == 0:
            return
        if (i % 32) < 12 or (i % 32) > 23:
            ksize=100
            roi = self.image[ROI[1]:ROI[1] + ROI[3], ROI[0]:ROI[0] + ROI[2]]
            roi = cv2.blur(roi, (ksize, ksize)) 
            self.image[ROI[1]:ROI[1] + ROI[3], ROI[0]:ROI[0] + ROI[2]] = roi 
            
            ROI.clear()
        
        self.image_save()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_dialog = MainDialog()
    main_dialog.show()
    app.exec_()
